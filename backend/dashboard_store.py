import json
import os
import shutil
from copy import deepcopy
from pathlib import Path
from typing import Any

from fastapi import HTTPException, status


BASE_DIR = Path(__file__).resolve().parent
APP_DIR = BASE_DIR.parent
SEED_DATA_PATH = APP_DIR / "src" / "data.json"
SEED_EXCEL_PATH = APP_DIR.parent / "sample_data.xlsx"
DASHBOARD_DATA_DIR = os.getenv("DASHBOARD_DATA_DIR", "").strip()
if DASHBOARD_DATA_DIR:
    DATA_DIR = Path(DASHBOARD_DATA_DIR)
    DATA_PATH = DATA_DIR / "data.json"
    EXCEL_PATH = DATA_DIR / "sample_data.xlsx"
else:
    DATA_PATH = SEED_DATA_PATH
    EXCEL_PATH = SEED_EXCEL_PATH
ALLOWED_DYNAMIC_TABLES = {"Page Settings"}


def ensure_dashboard_files() -> None:
    DATA_PATH.parent.mkdir(parents=True, exist_ok=True)

    if not DATA_PATH.exists() and SEED_DATA_PATH.exists():
        shutil.copyfile(SEED_DATA_PATH, DATA_PATH)

    if not EXCEL_PATH.exists() and SEED_EXCEL_PATH.exists():
        EXCEL_PATH.parent.mkdir(parents=True, exist_ok=True)
        shutil.copyfile(SEED_EXCEL_PATH, EXCEL_PATH)


def load_dashboard_data() -> dict[str, Any]:
    ensure_dashboard_files()

    if not DATA_PATH.exists():
        raise FileNotFoundError(f"Dashboard data file not found at: {DATA_PATH}")

    with DATA_PATH.open("r", encoding="utf-8") as data_file:
        return json.load(data_file)


def save_dashboard_data(data: dict[str, Any]) -> dict[str, Any]:
    ensure_dashboard_files()

    normalized = deepcopy(data)
    DATA_PATH.write_text(
        json.dumps(normalized, indent=2, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )
    save_excel_workbook(normalized)
    return normalized


def get_table(data: dict[str, Any], table_name: str) -> list[dict[str, Any]]:
    table = data.get(table_name)
    if not isinstance(table, list):
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Unknown dashboard table: {table_name}",
        )
    return table


def replace_table(table_name: str, rows: list[dict[str, Any]]) -> dict[str, Any]:
    data = load_dashboard_data()
    if table_name not in ALLOWED_DYNAMIC_TABLES:
        get_table(data, table_name)
    data[table_name] = rows
    return save_dashboard_data(data)


def save_excel_workbook(data: dict[str, Any]) -> None:
    try:
        from openpyxl import Workbook, load_workbook
    except ImportError as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Excel sync requires openpyxl. Install backend requirements before saving admin changes.",
        ) from exc

    EXCEL_PATH.parent.mkdir(parents=True, exist_ok=True)
    workbook = load_workbook(EXCEL_PATH) if EXCEL_PATH.exists() else Workbook()

    default_sheet = workbook.active
    if default_sheet.title == "Sheet" and "Sheet" not in data and len(workbook.sheetnames) == 1:
        workbook.remove(default_sheet)

    for table_name, rows in data.items():
        if not isinstance(rows, list):
            continue

        if table_name in workbook.sheetnames:
            sheet_index = workbook.sheetnames.index(table_name)
            existing_sheet = workbook[table_name]
            temp_title = f"__tmp_{sheet_index}"
            sheet = workbook.create_sheet(temp_title, sheet_index)
            workbook.remove(existing_sheet)
            sheet.title = table_name
        else:
            sheet = workbook.create_sheet(table_name)

        columns = get_columns(rows)
        if columns:
            sheet.append(columns)
            for row in rows:
                if isinstance(row, dict):
                    sheet.append([row.get(column) for column in columns])
                elif isinstance(row, list):
                    sheet.append(row)
                else:
                    sheet.append([row])

    workbook.save(EXCEL_PATH)


def get_excel_workbook_path() -> Path:
    data = load_dashboard_data()
    if not EXCEL_PATH.exists():
        save_excel_workbook(data)
    return EXCEL_PATH


def get_columns(rows: list[Any]) -> list[str]:
    columns: list[str] = []
    for row in rows:
        if isinstance(row, dict):
            for key in row.keys():
                if key not in columns:
                    columns.append(key)
        elif isinstance(row, list):
            width = len(row)
            columns.extend(str(index) for index in range(len(columns), width))
    return columns
